from openpyxl import *


wb = load_workbook('FinPom.xlsx')
sheet1 = wb.active


def excel():
    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 20
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 20
    sheet1.column_dimensions['G'].width = 20
    sheet1.column_dimensions['H'].width = 20
    sheet1.column_dimensions['I'].width = 20
    sheet1.column_dimensions['J'].width = 20
    sheet1.column_dimensions['K'].width = 20
    sheet1.column_dimensions['L'].width = 20
    sheet1.column_dimensions['M'].width = 20
    sheet1.column_dimensions['N'].width = 20
    sheet1.column_dimensions['O'].width = 20
    sheet1.column_dimensions['P'].width = 20
    sheet1.column_dimensions['Q'].width = 20


def wordfinder(ABC):
    for i in range(1, sheet1.max_row + 1):
        for j in range(1, sheet1.max_column + 1):
            if ABC == sheet1.cell(i, j).value:
                a = i
                return a


current_row = sheet1.max_row

if wordfinder(str(input("ID"))) == None:
    sheet1.cell(row=current_row + 1, column=1).value = input("ID")
    doh = str(input("Категория:"))
    if doh == "Соц.выплаты":
        sheet1.cell(row=current_row + 1, column=14).value = input("Доход:")
    elif doh == "зарплата":
        sheet1.cell(row=current_row + 1, column=15).value = input("Доход:")
    elif doh == "карманые деньги":
        sheet1.cell(row=current_row + 1, column=16).value = input("Доход:")
    else:
        sheet1.cell(row=current_row + 1, column=17).value = input("Доход:")
    ras = str(input("Категория:"))
    if ras == "развлечение":
        sheet1.cell(row=current_row + 1, column=6).value = input("Доход:")
    elif ras == "услуги":
        sheet1.cell(row=current_row + 1, column=7).value = input("Доход:")
    elif ras == "питание":
        sheet1.cell(row=current_row + 1, column=8).value = input("Доход:")
    elif ras == "медицина":
        sheet1.cell(row=current_row + 1, column=9).value = input("Доход:")
    elif ras == "транспорт":
        sheet1.cell(row=current_row + 1, column=10).value = input("Доход:")
    elif ras == "товары":
        sheet1.cell(row=current_row + 1, column=11).value = input("Доход:")
    else:
        sheet1.cell(row=current_row + 1, column=12).value = input("Доход:")

else:
    a = wordfinder(str(input("ID")))
    doh = str(input("Категория:"))
    if doh == "Соц.выплаты":
        sheet1.cell(row=a, column=14).value = input("Доход:")
    elif doh == "зарплата":
        sheet1.cell(row=a, column=15).value = input("Доход:")
    elif doh == "карманые деньги":
        sheet1.cell(row=a, column=16).value = input("Доход:")
    else:
        sheet1.cell(row=a, column=17).value = input("Доход:")
    ras = str(input("Категория:"))
    if ras == "развлечение":
        sheet1.cell(row=a, column=6).value = input("Доход:")
    elif ras == "услуги":
        sheet1.cell(row=a, column=7).value = input("Доход:")
    elif ras == "питание":
        sheet1.cell(row=a, column=8).value = input("Доход:")
    elif ras == "медицина":
        sheet1.cell(row=a, column=9).value = input("Доход:")
    elif ras == "транспорт":
        sheet1.cell(row=a, column=10).value = input("Доход:")
    elif ras == "товары":
        sheet1.cell(row=a, column=11).value = input("Доход:")
    else:
        sheet1.cell(row=a, column=12).value = input("Доход:")
excel()
wb.save('FinPom.xlsx')
