@bot.message_handler(state=MyStates.age, is_digit=True)
def ready_for_answer(message):
    """
    State 3. Will process when user's state is MyStates.age.
    """
    with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        msg = ("Ready, take a look:\n<b>"
               f"Name: {data['name']}\n"
               f"Surname: {data['surname']}\n"
               f"Age: {message.text}</b>")

        # here is this function
        func_to_use_next_in_code(data)

        bot.send_message(message.chat.id, msg, parse_mode="html")
    bot.delete_state(message.from_user.id, message.chat.id)


def func_to_use_next_in_code(data):
    # do stuff..
    print(data)





from openpyxl import *

wb = load_workbook('balances.xlsx')
sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10


current_row = sheet.max_row
current_column = sheet.max_column


sheet.cell(row=current_row + 1, column=1).value = input()
sheet.cell(row=current_row + 1, column=2).value = input()
wb.save('balances.xlsx')
if __name__ == "__main__":
    excel()
wb.save('balances.xlsx')
