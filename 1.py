from openpyxl import *
import datetime

wb = load_workbook('FinPom2.xlsx')

sheet2 = wb.worksheets[1]
name = ""
nal = 100
beznal = 500
nal2 = 500
beznal2 = 100


def doch1():
    sheet2.cell(row=current_row + 1, column=5).value = datetime.date.today()
    sheet2.cell(row=current_row + 1, column=1).value = name
    sheet2.cell(row=current_row + 1, column=3).value = "Наличныe"
    sheet2.cell(row=current_row + 1, column=2).value = nal
    sheet2.cell(row=current_row + 1, column=4).value = "Доход"


def doch2():
    sheet2.cell(row=current_row + 1, column=5).value = datetime.date.today()
    sheet2.cell(row=current_row + 1, column=1).value = name
    sheet2.cell(row=current_row + 1, column=3).value = "Безнал"
    sheet2.cell(row=current_row + 1, column=2).value = beznal
    sheet2.cell(row=current_row + 1, column=4).value = "Доход"


def rash1():
    sheet2.cell(row=current_row + 1, column=5).value = datetime.date.today()
    sheet2.cell(row=current_row + 1, column=1).value = name
    sheet2.cell(row=current_row + 1, column=3).value = "Наличныe"
    sheet2.cell(row=current_row + 1, column=2).value = nal2
    sheet2.cell(row=current_row + 1, column=4).value = "Расход"


def rash2():
    sheet2.cell(row=current_row + 1, column=5).value = datetime.date.today()
    sheet2.cell(row=current_row + 1, column=1).value = name
    sheet2.cell(row=current_row + 1, column=3).value = "Безнал"
    sheet2.cell(row=current_row + 1, column=2).value = beznal2
    sheet2.cell(row=current_row + 1, column=4).value = "Расход"

current_row = sheet2.max_row

doch1()

wb.save('FinPom2.xlsx')
