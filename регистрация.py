from openpyxl import *

wb = load_workbook('FinPom2.xlsx')
sheet1 = wb.worksheets[0]
name = input("ID:")

current_row = sheet1.max_row

def wordfinder(ABC):
    for i in range(1, sheet1.max_row + 1):
        for j in range(1, sheet1.max_column + 1):
            if ABC == sheet1.cell(i, j).value:
                a = i
                return a


if wordfinder(str(name)) == None:
    sheet1.cell(row=current_row + 1, column=1).value = name
else:
    a = wordfinder(str(name))
    sheet1.cell(row=a, column=1).value = name
wb.save('FinPom2.xlsx')