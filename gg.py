from openpyxl import Workbook

wb = Workbook()

a = 'dsvkjdnvkjdkdnfdnvfd'
n = '0345'

ws1 = wb.create_sheet("reg", 0)



def vhod():
    for row in range(1, 3361):
        table_name = str(ws1[row][0].value)
        table_password = str(ws1[row][1].value)
        if a == table_name and n == table_password:
            print()
        else:
            print("Вы зарегистрированы!")
#           \/
#          вход

def register():
    for row in range(1, 3361):
        table_name = str(ws1[row][0].value)
        if a == table_name:
            print("Такой логи уже есть")
        else:
            print("Вы зарегистрированы!")
#           \/
#       регистрация





ws2 = wb.create_sheet("ras", 1)
ws3 = wb.create_sheet("doh", 2)

wb.save('FinPom.xlsx')
